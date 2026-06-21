from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# Colour palette
COLOUR_HEADER_BG = "1F4E79"
COLOUR_HEADER_FG = "FFFFFF"
COLOUR_MASTERED = "C6EFCE"
COLOUR_IN_PROGRESS = "FFEB9C"
COLOUR_UNMATCHED = "FFC7CE"
COLOUR_ALT_ROW = "F2F2F2"

OUTPUT_PATH = Path("data/ra_collection.xlsx")

# Console sheet columns: (header label, DataFrame column, width)
CONSOLE_COLUMNS = [
    ("Filename", "filename", 30),
    ("RA Title", "ra_title", 35),
    ("Matched", "matched", 10),
    ("Earned", "earned", 10),
    ("Total", "total", 10),
    ("Completion %", "completion_pct", 14),
    ("Mastered", "is_mastered", 12),
    ("Status", "status", 22),
    ("Console", "console", 12),
]

# Unmatched sheet columns: (header label, DataFrame column, width)
UNMATCHED_COLUMNS = [
    ("Original Filename", "filename", 30),
    ("Console", "console", 12),
    ("Suggested Title", "suggested_title", 35),
    ("Expected Dump Name", "suggested_filename", 45),
    ("Expected MD5", "suggested_md5", 35),
    ("Patch URL", "patch_url", 30),
]


def header_font(bold: bool = True) -> Font:
    return Font(name="Arial", bold=bold, color=COLOUR_HEADER_FG, size=11)


def body_font(bold: bool = False) -> Font:
    return Font(name="Arial", bold=bold, size=10)


def _header_fill() -> PatternFill:
    return PatternFill("solid", start_color=COLOUR_HEADER_BG)


def row_fill(colour: str) -> PatternFill:
    return PatternFill("solid", start_color=colour)


def write_header_row(ws, headers: list[str]) -> None:
    for col_idx, label in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = header_font()
        cell.fill = _header_fill()
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 20


def set_column_widths(ws, columns: list[tuple]) -> None:
    for col_idx, (_, _, width) in enumerate(columns, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _row_colour(status: str) -> str | None:
    if status == "Mastered 🏆":
        return COLOUR_MASTERED
    if str(status).startswith("In Progress"):
        return COLOUR_IN_PROGRESS
    if status in ("Unknown/Unlinked", "Unmatched"):
        return COLOUR_UNMATCHED
    return None


def _write_console_sheet(wb: Workbook, console_name: str, df: pd.DataFrame) -> None:
    ws = wb.create_sheet(title=console_name[:31])  # Excel sheet name limit

    headers = [col[0] for col in CONSOLE_COLUMNS]
    write_header_row(ws, headers)
    set_column_widths(ws, CONSOLE_COLUMNS)
    ws.freeze_panes = "A2"

    for row_idx, (_, row) in enumerate(df.iterrows(), start=2):
        fill_colour = _row_colour(str(row.get("status", "")))
        alt_colour = COLOUR_ALT_ROW if row_idx % 2 == 0 else None

        for col_idx, (_, col_key, _) in enumerate(CONSOLE_COLUMNS, start=1):
            value = row.get(col_key, "")
            # Clean up pandas NA/NaN for display
            if pd.isna(value):
                value = ""
            elif isinstance(value, bool):
                value = "Yes" if value else "No"

            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = body_font()
            cell.alignment = Alignment(vertical="center")

            if fill_colour:
                cell.fill = row_fill(fill_colour)
            elif alt_colour:
                cell.fill = row_fill(alt_colour)

    ws.auto_filter.ref = f"A1:{get_column_letter(len(CONSOLE_COLUMNS))}1"


def _write_summary_sheet(
    wb: Workbook,
    df: pd.DataFrame,
    user_summary: dict | None,
) -> None:
    ws = wb.create_sheet(title="Summary", index=0)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 20

    title_font = Font(name="Arial", bold=True, size=14, color=COLOUR_HEADER_BG)
    label_font = Font(name="Arial", bold=True, size=11)
    value_font = Font(name="Arial", size=11)
    section_fill = row_fill("DCE6F1")

    def write_section_header(row: int, label: str) -> None:
        cell = ws.cell(row=row, column=1, value=label)
        cell.font = label_font
        cell.fill = section_fill
        ws.merge_cells(f"A{row}:B{row}")

    def write_row(row: int, label: str, value) -> None:
        lc = ws.cell(row=row, column=1, value=label)
        lc.font = body_font(bold=True)
        vc = ws.cell(row=row, column=2, value=value)
        vc.font = value_font

    ws.cell(row=1, column=1, value="RA ROM Manager - Collection Summary").font = title_font
    ws.merge_cells("A1:B1")
    ws.row_dimensions[1].height = 28

    # User profile section
    write_section_header(3, "👤 RA Profile")
    if user_summary:
        write_row(4, "Points", user_summary.get("points", "-"))
        write_row(5, "Softcore Points", user_summary.get("softcore_points", "-"))
        write_row(6, "Global Rank", user_summary.get("rank", "-"))
        write_row(7, "Games Played", user_summary.get("games_played", "-"))
    else:
        write_row(4, "Profile data", "unavailable")

    # Collection stats section - use Excel formulas referencing console sheets
    total_roms = len(df)
    matched = int(df["matched"].sum()) if "matched" in df.columns else 0
    mastered = int(df["is_mastered"].sum()) if "is_mastered" in df.columns else 0
    in_progress = (
        int(df["status"].str.startswith("In Progress").sum()) if "status" in df.columns else 0
    )
    unplayed = int((df["status"] == "Unplayed").sum()) if "status" in df.columns else 0
    unmatched = total_roms - matched

    write_section_header(9, "🎮 Collection")
    write_row(10, "Total ROMs", total_roms)
    write_row(11, "Matched to RA", matched)
    write_row(12, "Unmatched", unmatched)
    write_row(13, "Match Rate", f"{(matched / total_roms * 100):.1f}%" if total_roms else "-")

    write_section_header(15, "🏆 Progress")
    write_row(16, "Mastered", mastered)
    write_row(17, "In Progress", in_progress)
    write_row(18, "Unplayed", unplayed)

    # Per-console breakdown
    write_section_header(20, "📋 By Console")
    ws.cell(row=21, column=1, value="Console").font = body_font(bold=True)
    ws.cell(row=21, column=2, value="ROMs").font = body_font(bold=True)

    consoles = df["console"].str.upper().value_counts() if "console" in df.columns else {}
    for offset, (console, count) in enumerate(consoles.items()):
        write_row(22 + offset, console, count)


def _write_want_to_play_sheet(
    wb: Workbook,
    collection_df: pd.DataFrame | None = None,
    client=None,
) -> None:
    want_to_play_path = Path("data/want_to_play.csv")
    ws = wb.create_sheet(title="Want to Play")

    headers = [
        "RA Game ID", "Title", "Console", "Notes", "Added Date",
        "Owned", "Achievements", "Earned", "Completion %", "Status",
    ]
    col_widths = [12, 35, 16, 30, 14, 8, 14, 10, 14, 22]

    write_header_row(ws, headers)
    for col_idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = "A2"

    owned_ids: set = set()
    if collection_df is not None and "ra_game_id" in collection_df.columns:
        owned_ids = set(
            collection_df.loc[collection_df["matched"], "ra_game_id"].dropna().astype(int)
        )

    if want_to_play_path.exists():
        wtp_df = pd.read_csv(want_to_play_path)
        for row_idx, (_, row) in enumerate(wtp_df.iterrows(), start=2):
            game_id_raw = row.get("ra_game_id")
            game_id = int(game_id_raw) if pd.notna(game_id_raw) else None

            owned = "Yes" if game_id in owned_ids else "No"
            achievements = earned = completion = status = ""

            if client and game_id:
                try:
                    progress = client.get_user_progress(game_id)
                    achievements = progress.get("total", "")
                    earned = progress.get("earned", "")
                    total = progress.get("total", 0)
                    completion = f"{progress['earned'] / total * 100:.1f}%" if total else "-"
                    status = (
                        "Mastered 🏆" if progress.get("is_mastered")
                        else f"In Progress ({completion})" if earned
                        else "Unplayed"
                    )
                except Exception:
                    pass

            values = [
                game_id_raw if pd.notna(game_id_raw) else "",
                row.get("title", ""), row.get("console", ""),
                row.get("notes", ""), row.get("added_date", ""),
                owned, achievements, earned, completion, status,
            ]
            for col_idx, value in enumerate(values, start=1):
                if pd.isna(value):
                    value = ""
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = body_font()
                cell.alignment = Alignment(vertical="center")
    else:
        ws.cell(row=2, column=1, value="No want_to_play.csv found in data/").font = Font(
            name="Arial", italic=True, color="999999"
        )


def _write_unmatched_sheet(wb: Workbook, df: pd.DataFrame) -> None:
    ws = wb.create_sheet(title="Unmatched ROMs")

    headers = [col[0] for col in UNMATCHED_COLUMNS]
    write_header_row(ws, headers)
    set_column_widths(ws, UNMATCHED_COLUMNS)
    ws.freeze_panes = "A2"

    # Filter only unmatched ROMs
    unmatched_df = df[~df["matched"]]

    if unmatched_df.empty:
        ws.cell(row=2, column=1, value="All ROMs matched perfectly!").font = Font(
            name="Arial", italic=True, color="999999"
        )
        return

    for row_idx, (_, row) in enumerate(unmatched_df.iterrows(), start=2):
        alt_colour = COLOUR_ALT_ROW if row_idx % 2 == 0 else None

        for col_idx, (_, col_key, _) in enumerate(UNMATCHED_COLUMNS, start=1):
            value = row.get(col_key, "")
            if pd.isna(value):
                value = ""

            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = body_font()
            cell.alignment = Alignment(vertical="center")

            # Make Patch URL a clickable hyperlink if it exists
            if col_key == "patch_url" and value:
                cell.hyperlink = value
                cell.font = Font(name="Arial", size=10, color="0563C1", underline="single")

            if alt_colour:
                cell.fill = row_fill(alt_colour)


def export(
    df: pd.DataFrame,
    user_summary: dict | None = None,
    output_path: Path | str | None = None,
    client=None,
) -> Path:
    """
    Builds and saves the full Excel workbook to data/ra_collection.xlsx.

    Sheets:
        Summary      - user stats + collection overview (always first)
        <Console>    - one sheet per console, with conditional formatting
        Want to Play - sourced from data/want_to_play.csv if present

    Returns the path to the saved file.
    """
    # Use the provided path, or default to the original
    target_path = Path(output_path) if output_path else OUTPUT_PATH
    target_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    # Remove the default empty sheet
    wb.remove(wb.active)

    # Summary first
    _write_summary_sheet(wb, df, user_summary)

    # One sheet per console
    consoles = df["console"].str.upper().unique() if "console" in df.columns else []
    for console in sorted(consoles):
        console_df = df[df["console"].str.upper() == console].copy()
        _write_console_sheet(wb, console, console_df)

    # Write the unmatched sheet
    _write_unmatched_sheet(wb, df)

    # Write the want to play sheet
    _write_want_to_play_sheet(wb, collection_df=df, client=client)

    wb.save(target_path)
    return target_path
