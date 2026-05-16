import unicodedata
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font

from .config import CONSOLES
from .exporter import COLOUR_ALT_ROW, body_font, row_fill, set_column_widths, write_header_row


def remove_accents(text: str) -> str:
    """Removes accents to make searching easier (e.g., Pokémon -> Pokemon)"""
    return unicodedata.normalize('NFKD', text).encode('ASCII', 'ignore').decode('utf-8')

def run_franchise_report(keyword: str, client) -> None:
    print(f"\n🔍 Searching RetroAchievements database for '{keyword}'...")

    matched_games =[]
    search_term = remove_accents(keyword.lower())

    # 1. Search all consoles
    for console_id, console_name in CONSOLES.items():
        try:
            game_list = client.get_console_game_hashes(console_id)
            for game in game_list:
                title = game.get("Title", "")

                # Strip accents from the RA title for the comparison
                clean_title = remove_accents(title.lower())

                if search_term in clean_title:
                    game_id = game.get("ID")
                    # Prevent duplicates
                    if not any(g["ra_game_id"] == game_id for g in matched_games):
                        matched_games.append({
                            "ra_game_id": game_id,
                            "ra_title": title,
                            "console": console_name
                        })
        except Exception as e:
            print(f"   ⚠️  Skipping {console_name}: {e}")
            continue

    if not matched_games:
        print(f"   ⚠️  No games found matching '{keyword}'.")
        return

    print(f"🎯 Found {len(matched_games)} matching games! Fetching your progress...")

    # 2. Fetch Progress
    results =[]
    for idx, game in enumerate(matched_games, start=1):
        print(f"[{idx}/{len(matched_games)}] Checking {game['ra_title']}...")
        try:
            prog = client.get_user_progress(game["ra_game_id"])

            # Skip games that don't actually have achievements
            if prog["total"] == 0:
                continue

            game.update({
                "earned": prog["earned"],
                "total": prog["total"],
                "points_earned": prog.get("points_earned", 0),
                "points_total": prog.get("points_total", 0),
                "completion_pct": (
                    round((prog["earned"] / prog["total"]) * 100, 1)
                    if prog["total"] > 0 else 0.0
                ),
                "is_mastered": prog["is_mastered"],
                "status": "Played" if prog["earned"] > 0 else "To Play"
            })
            results.append(game)
        except Exception:
            print(f"   ⚠️  Could not fetch progress for {game['ra_title']}")

    # SAFETY CHECK: Make sure we actually have data before exporting
    if not results:
        print("\n   ⚠️  Found games, but none of them had active achievements. Nothing to export!")
        return

    df = pd.DataFrame(results)

    # 3. Export custom Excel report
    _export_franchise_excel(keyword, df)


def _export_franchise_excel(keyword: str, df: pd.DataFrame):
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_keyword = keyword.replace(" ", "_")
    output_path = Path("data") / f"franchise_{safe_keyword}_{timestamp}.xlsx"
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)

    # Split DataFrames
    played_df = df[df["status"] == "Played"].sort_values("completion_pct", ascending=False)
    toplay_df = df[df["status"] == "To Play"].sort_values("console")

    # Sheet 1: Stats & Overview
    ws_stats = wb.create_sheet(title="Franchise Stats")
    ws_stats.column_dimensions["A"].width = 30
    ws_stats.column_dimensions["B"].width = 15

    title_cell = ws_stats.cell(row=1, column=1, value=f"Franchise Report: {keyword.upper()}")
    title_cell.font = Font(bold=True, size=14)

    total_games = len(df)
    games_played = len(played_df)
    mastered = played_df["is_mastered"].sum() if not played_df.empty else 0
    total_achievements = df["total"].sum()
    earned_achievements = played_df["earned"].sum() if not played_df.empty else 0
    total_points = played_df["points_earned"].sum() if not played_df.empty else 0

    stats =[
        ("Total Games with Achievements", total_games),
        ("Games Played", games_played),
        ("Games Mastered 🏆", mastered),
        ("Achievements Earned", f"{earned_achievements} / {total_achievements}"),
        ("Total Franchise Points", total_points)
    ]

    for row_idx, (label, val) in enumerate(stats, start=3):
        ws_stats.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
        ws_stats.cell(row=row_idx, column=2, value=val)

    # Helper function for data sheets
    def write_data_sheet(title, data_df):
        if data_df.empty:
            return  # Skip creating the sheet if there's no data for it

        ws = wb.create_sheet(title=title)
        cols =[("Title", "ra_title", 45), ("Console", "console", 15),
                ("Earned", "earned", 10), ("Total", "total", 10),
                ("%", "completion_pct", 10), ("Mastered", "is_mastered", 12)]

        write_header_row(ws, [c[0] for c in cols])
        set_column_widths(ws, cols)

        for row_idx, (_, row) in enumerate(data_df.iterrows(), start=2):
            alt_colour = COLOUR_ALT_ROW if row_idx % 2 == 0 else None

            for col_idx, (_, key, _) in enumerate(cols, start=1):
                val = row.get(key, "")

                if key == "is_mastered":
                    val = "Yes" if val else "No"

                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.font = body_font()

                if alt_colour:
                    cell.fill = row_fill(alt_colour)

    write_data_sheet("Played", played_df)
    write_data_sheet("To Play", toplay_df)

    wb.save(output_path)
    print(f"\n💾 Saved Franchise Report to {output_path}")
