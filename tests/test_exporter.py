from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

from src.ra_manager.exporter import export


@pytest.fixture
def sample_df() -> pd.DataFrame:
    return pd.DataFrame(
        {
            "filename": ["rayman.gba", "pokemon.gba", "metroid.gba", "unknown.gba"],
            "console": ["gba", "gba", "gba", "gba"],
            "md5": [
                "fb20d6009c7400f37581f81ae5b1e917",
                "dfc6fdf38b3c277b6f176cd7c25712c8",
                "6dab0ac88b4e438092c2a90338e51a1b",
                "000000000000000000000000notareal",
            ],
            "ra_title": [
                "Rayman Advance",
                "Pokémon LeafGreen Version",
                "Metroid Fusion",
                "Unknown/Unlinked",
            ],
            "ra_game_id": [1141, 1448, 2200, None],
            "matched": [True, True, True, False],
            "earned": [15, 73, 0, None],
            "total": [50, 73, 40, None],
            "completion_pct": [30.0, 100.0, 0.0, None],
            "is_mastered": [False, True, False, False],
            "status": ["In Progress (30.0%)", "Mastered 🏆", "Unplayed", "Unmatched"],
        }
    )


@pytest.fixture
def sample_summary() -> dict:
    return {
        "points": 4200,
        "softcore_points": 300,
        "rank": 18500,
        "games_played": 42,
    }


@pytest.fixture
def output_path(tmp_path, monkeypatch) -> Path:
    """Redirect OUTPUT_PATH to tmp dir for all export tests."""
    import src.ra_manager.exporter as exporter_module

    path = tmp_path / "ra_collection.xlsx"
    monkeypatch.setattr(exporter_module, "OUTPUT_PATH", path)
    return path


class TestExportFileCreation:
    def test_creates_xlsx_file(self, sample_df, sample_summary, output_path):
        export(sample_df, sample_summary)
        assert output_path.exists()

    def test_returns_output_path(self, sample_df, sample_summary, output_path):
        result = export(sample_df, sample_summary)
        assert result == output_path

    def test_creates_data_dir_if_missing(self, sample_df, tmp_path, monkeypatch):
        import src.ra_manager.exporter as exporter_module

        nested = tmp_path / "nested" / "data" / "ra_collection.xlsx"
        monkeypatch.setattr(exporter_module, "OUTPUT_PATH", nested)
        export(sample_df)
        assert nested.exists()


class TestSheetStructure:
    def test_summary_sheet_is_first(self, sample_df, sample_summary, output_path):
        export(sample_df, sample_summary)
        wb = load_workbook(output_path)
        assert wb.sheetnames[0] == "Summary"

    def test_console_sheet_created_for_gba(self, sample_df, sample_summary, output_path):
        export(sample_df, sample_summary)
        wb = load_workbook(output_path)
        assert "GBA" in wb.sheetnames

    def test_want_to_play_sheet_exists(self, sample_df, sample_summary, output_path):
        export(sample_df, sample_summary)
        wb = load_workbook(output_path)
        assert "Want to Play" in wb.sheetnames

    def test_want_to_play_is_last_sheet(self, sample_df, sample_summary, output_path):
        export(sample_df, sample_summary)
        wb = load_workbook(output_path)
        assert wb.sheetnames[-1] == "Want to Play"


class TestConsoleSheet:
    def test_header_row_contains_expected_columns(self, sample_df, output_path):
        export(sample_df)
        wb = load_workbook(output_path)
        ws = wb["GBA"]
        headers = [ws.cell(row=1, column=c).value for c in range(1, 10)]
        assert "Filename" in headers
        assert "RA Title" in headers
        assert "Status" in headers

    def test_all_roms_written_to_sheet(self, sample_df, output_path):
        export(sample_df)
        wb = load_workbook(output_path)
        ws = wb["GBA"]
        # 4 ROMs + 1 header row
        assert ws.max_row == 5

    def test_rom_filename_in_first_data_row(self, sample_df, output_path):
        export(sample_df)
        wb = load_workbook(output_path)
        ws = wb["GBA"]
        filenames = [ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)]
        assert "rayman.gba" in filenames


class TestSummarySheet:
    def test_user_points_written(self, sample_df, sample_summary, output_path):
        export(sample_df, sample_summary)
        wb = load_workbook(output_path)
        ws = wb["Summary"]
        values = [ws.cell(row=r, column=2).value for r in range(1, 20)]
        assert 4200 in values

    def test_total_roms_written(self, sample_df, sample_summary, output_path):
        export(sample_df, sample_summary)
        wb = load_workbook(output_path)
        ws = wb["Summary"]
        values = [ws.cell(row=r, column=2).value for r in range(1, 25)]
        assert 4 in values  # 4 total ROMs in sample_df

    def test_summary_works_without_user_summary(self, sample_df, output_path):
        export(sample_df, user_summary=None)
        wb = load_workbook(output_path)
        assert "Summary" in wb.sheetnames


class TestUnmatchedSheet:
    def test_unmatched_sheet_created_if_data_exists(self, sample_df, output_path):
        # Modify sample_df to have M5 suggestion columns
        sample_df["suggested_title"] = ["", "", "", "Suggested Unknown Game"]
        sample_df["suggested_filename"] = ["", "", "", "Unknown (USA).gba"]
        sample_df["suggested_md5"] = ["", "", "", "12345abcdef"]
        sample_df["patch_url"] = ["", "", "", "http://patch.com"]

        export(sample_df)
        wb = load_workbook(output_path)

        assert "Unmatched ROMs" in wb.sheetnames
        ws = wb["Unmatched ROMs"]

        # Row 1 is header, Row 2 should be the unmatched ROM
        assert ws.cell(row=2, column=3).value == "Suggested Unknown Game"

    def test_all_matched_shows_placeholder_message(self, sample_df, output_path):
        all_matched = sample_df[sample_df["matched"]].copy().reset_index(drop=True)
        export(all_matched)
        wb = load_workbook(output_path)
        ws = wb["Unmatched ROMs"]
        assert ws.cell(row=2, column=1).value == "All ROMs matched perfectly!"

    def test_patch_url_written_as_hyperlink(self, sample_df, output_path):
        sample_df["suggested_title"] = ["", "", "", "Some Title"]
        sample_df["suggested_filename"] = ["", "", "", "Some (USA).gba"]
        sample_df["suggested_md5"] = ["", "", "", "abc123"]
        sample_df["patch_url"] = ["", "", "", "https://example.com/patch"]
        export(sample_df)
        wb = load_workbook(output_path)
        ws = wb["Unmatched ROMs"]
        patch_cell = ws.cell(row=2, column=6)
        assert patch_cell.hyperlink is not None
        assert "example.com" in patch_cell.hyperlink.target


WTP_CSV_RAYMAN = (
    "ra_game_id,title,console,notes,added_date,owned\n"
    "1141,Rayman Advance,gba,,2026-01-01,\n"
)
WTP_CSV_UNKNOWN = (
    "ra_game_id,title,console,notes,added_date,owned\n"
    "9999,Unknown Game,gba,,2026-01-01,\n"
)


class TestWantToPlayEnrichment:
    def _make_client(self, progress):
        client = type("Client", (), {})()
        client.get_user_progress = lambda game_id, **kw: progress
        return client

    def _setup_wtp_csv(self, tmp_path, content):
        data_dir = tmp_path / "data"
        data_dir.mkdir()
        (data_dir / "want_to_play.csv").write_text(content)

    def test_progress_columns_written_when_client_provided(
        self, sample_df, output_path, tmp_path, monkeypatch
    ):
        import src.ra_manager.exporter as exporter_module
        self._setup_wtp_csv(tmp_path, WTP_CSV_RAYMAN)
        monkeypatch.setattr(exporter_module, "OUTPUT_PATH", output_path)
        monkeypatch.chdir(tmp_path)

        client = self._make_client({"earned": 10, "total": 50, "is_mastered": False})
        export(sample_df, client=client)
        wb = load_workbook(output_path)
        ws = wb["Want to Play"]
        headers = [ws.cell(row=1, column=c).value for c in range(1, 11)]
        assert "Achievements" in headers
        assert "Earned" in headers

    def test_owned_flag_set_for_matched_game(
        self, sample_df, output_path, tmp_path, monkeypatch
    ):
        import src.ra_manager.exporter as exporter_module
        self._setup_wtp_csv(tmp_path, WTP_CSV_RAYMAN)
        monkeypatch.setattr(exporter_module, "OUTPUT_PATH", output_path)
        monkeypatch.chdir(tmp_path)

        client = self._make_client({"earned": 0, "total": 50, "is_mastered": False})
        export(sample_df, client=client)
        wb = load_workbook(output_path)
        ws = wb["Want to Play"]
        assert ws.cell(row=2, column=6).value == "Yes"

    def test_not_owned_for_unmatched_game(
        self, sample_df, output_path, tmp_path, monkeypatch
    ):
        import src.ra_manager.exporter as exporter_module
        self._setup_wtp_csv(tmp_path, WTP_CSV_UNKNOWN)
        monkeypatch.setattr(exporter_module, "OUTPUT_PATH", output_path)
        monkeypatch.chdir(tmp_path)

        client = self._make_client({"earned": 0, "total": 10, "is_mastered": False})
        export(sample_df, client=client)
        wb = load_workbook(output_path)
        ws = wb["Want to Play"]
        assert ws.cell(row=2, column=6).value == "No"

    def test_no_client_does_not_crash(
        self, sample_df, output_path, tmp_path, monkeypatch
    ):
        import src.ra_manager.exporter as exporter_module
        self._setup_wtp_csv(tmp_path, WTP_CSV_RAYMAN)
        monkeypatch.setattr(exporter_module, "OUTPUT_PATH", output_path)
        monkeypatch.chdir(tmp_path)

        export(sample_df)  # no client
        wb = load_workbook(output_path)
        assert "Want to Play" in wb.sheetnames
